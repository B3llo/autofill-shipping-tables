#  Script for transfering Rodonaves Price Sheets to another model table
#  that will create a SQL query for inserting into the database
#
#  Version Date: 28/01/2022


# Libraries imports 

from openpyxl import *
import pandas as pd 
import math
import re


##########     Importing and formatting the sheets     ##########


#  Insert here the names of the sheets to be imported
#  In order to work, the files must be in the same directory as this script

Price_Table_Name = "ASTEC.xlsx"
Model_Table_Name = "Planilha Geral para fretes.xlsx"

wb2 = load_workbook(filename = Price_Table_Name)
wb = load_workbook(filename = Model_Table_Name)


#  Change to the correct worksheet names of your files 

Model_WorkSheet = "frete"
Price_Table_WorkSheet = "Table 1"

model = wb[Model_WorkSheet]
prices = wb2[Price_Table_WorkSheet]

itemsToBeDeleted = 6  #  Number of the first elements that will be deleted from the Rodonaves workbook.

prices.delete_rows(0, itemsToBeDeleted)


#  New File Name

Updated_Price_Table = "MinhasBola.xlsx"
wb2.save(filename = Updated_Price_Table)

priceTable = pd.read_excel(Updated_Price_Table)
wb2 = load_workbook(filename = Updated_Price_Table)
prices = wb2["Table 1"]

##########     Counting the number of columns     ##########


avg = 0 # Variable to storage the average price for each state

weightIndex = 3   # Column for packages with weight less than 10 kg start at index 3 
maxWeightIndex =  -3   # weightIndex plus the number of other weight columns to include


for col in prices.iter_cols():
    maxWeightIndex += 1



##########     Calculating the average price for each state and weight group     ##########


weightGroup = []

while (maxWeightIndex >= weightIndex):

    priceGroup = []
    avgGroup = []
    
    for col in priceTable.itertuples():
        avg = 0
        NaN = math.isnan(col[weightIndex])

        if (col[0] == 1 and col[weightIndex] == NaN):
            break

        if (not NaN):
            priceGroup.append(col[weightIndex])

        elif(len(priceGroup) > 0):
            for n in priceGroup:
                avg += n

            if len(priceGroup) > 0:
                avg /= len(priceGroup)
                avgGroup.append(round(avg, 2))

            priceGroup = []

    if (len(avgGroup) > 0):
        weightGroup.append(avgGroup)
    
    weightIndex += 1 

for k in weightGroup:
    print(k)

# Assigning each average price to the corresponding model table


States = ["PR", "SC", "RS", "SP", "MG", "DF", "GO", "RJ", "ES", "MS", "MG", "RO", "AC", "PA"];

startCol = 3  # First column that will be modified
statesCols = -3 # Counting how many columns are until the end of the document

for col in model.iter_cols():
    statesCols += 1


##########     Modifying the values into the model table and saving the new document    ##########

for i in range(startCol, statesCols):

    for j in range(len(States)):
        match = re.search(f"{States[j]}", model.cell(row = 1, column = i).value)

        if(match):
            stateIndex = States.index(States[j])
            print(States[j], i, " [OK]")

            for n in range(2, 8):
                if(stateIndex <=12):
                    model.cell(row = n, column = i).value = weightGroup[(n - 2)][stateIndex]


wb.save(filename = Updated_Price_Table)